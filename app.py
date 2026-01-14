import io
import time
from collections import defaultdict
from typing import Dict, Tuple

import pandas as pd
import streamlit as st

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string


# =========================
# Helpers
# =========================
def safe_sheet_name(name: str) -> str:
    """Excel: max 31 chars, cannot contain : \\ / ? * [ ]"""
    banned = [":", "\\", "/", "?", "*", "[", "]"]
    for ch in banned:
        name = name.replace(ch, "")
    name = (name or "").strip() or "лист"
    return name[:31]


def split_prefix_suffix4(sheet_name: str) -> Tuple[str, str]:
    """prefix + last 4 chars (lowercased)"""
    if len(sheet_name) < 4:
        return sheet_name, ""
    return sheet_name[:-4], sheet_name[-4:].lower()


def split_prefix_suffix2(sheet_name: str) -> Tuple[str, str]:
    """prefix + last 2 chars (lowercased)"""
    if len(sheet_name) < 2:
        return sheet_name, ""
    return sheet_name[:-2], sheet_name[-2:].lower()


def normalize_prefix(prefix: str) -> str:
    return (prefix or "").strip()


# =========================
# CODE 1 (Saldo) — multi-company by prefix + ####
# Finds sheets where last 4 chars are 1210/1710/3310/3510
# Creates output sheet per prefix: "<prefix>сальд" or "сальд" if no prefix
# =========================
def run_code_1(file_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(file_bytes))
    xls = pd.ExcelFile(io.BytesIO(file_bytes))

    target_suffixes = {"1210": 6, "1710": 6, "3310": 7, "3510": 7}

    prefix_to_sheets: Dict[str, Dict[str, str]] = defaultdict(dict)
    for sh in xls.sheet_names:
        prefix, suf = split_prefix_suffix4(sh)
        if suf in target_suffixes:
            prefix = normalize_prefix(prefix)
            if suf not in prefix_to_sheets[prefix]:
                prefix_to_sheets[prefix][suf] = sh

    if not prefix_to_sheets:
        raise ValueError("Код 1: не найдено листов, заканчивающихся на 1210/1710/3310/3510.")

    for prefix, suf_map in prefix_to_sheets.items():
        sheet_data = {}

        for suf, col_idx in target_suffixes.items():
            if suf not in suf_map:
                continue

            df = pd.read_excel(xls, sheet_name=suf_map[suf], header=0)
            if df.shape[1] <= col_idx:
                continue

            temp = df.iloc[:, [0, col_idx]].copy()
            temp.columns = ["Контрагент", "value"]
            temp["value"] = pd.to_numeric(temp["value"], errors="coerce").fillna(0)
            temp = temp[temp["value"] != 0]
            temp["Контрагент"] = temp["Контрагент"].astype(str).str.strip()
            temp = temp[temp["Контрагент"] != ""]
            temp = temp[~temp["Контрагент"].isin(["1210", "1710", "3310", "3510"])]
            temp = temp[~temp["Контрагент"].str.lower().str.startswith("итого")]

            if temp.empty:
                continue

            sheet_data[suf] = temp.groupby("Контрагент")["value"].sum()

        s1210 = sheet_data.get("1210", pd.Series(dtype=float))
        s3510 = sheet_data.get("3510", pd.Series(dtype=float))
        s1710 = sheet_data.get("1710", pd.Series(dtype=float))
        s3310 = sheet_data.get("3310", pd.Series(dtype=float))

        cust_set = set(s1210.index).union(set(s3510.index))
        supp_set = set(s1710.index).union(set(s3310.index))
        all_set = cust_set.union(supp_set)

        if not cust_set and not supp_set:
            continue

        if cust_set:
            df_cust = pd.DataFrame(sorted(cust_set), columns=["Контрагент"])
            df_cust["1210"] = df_cust["Контрагент"].map(s1210).fillna(0) / 1000
            df_cust["3510"] = df_cust["Контрагент"].map(s3510).fillna(0) / 1000
            df_cust["сальдо заказчики"] = df_cust["1210"] - df_cust["3510"]
            df_cust = df_cust.sort_values(by="сальдо заказчики", ascending=False).reset_index(drop=True)
        else:
            df_cust = pd.DataFrame(columns=["Контрагент", "1210", "3510", "сальдо заказчики"])

        if supp_set:
            df_supp = pd.DataFrame(sorted(supp_set), columns=["Контрагент"])
            df_supp["1710"] = df_supp["Контрагент"].map(s1710).fillna(0) / 1000
            df_supp["3310"] = df_supp["Контрагент"].map(s3310).fillna(0) / 1000
            df_supp["сальдо поставщики"] = df_supp["1710"] - df_supp["3310"]
            df_supp = df_supp.sort_values(by="сальдо поставщики", ascending=False).reset_index(drop=True)
        else:
            df_supp = pd.DataFrame(columns=["Контрагент", "1710", "3310", "сальдо поставщики"])

        if all_set:
            df_total = pd.DataFrame(sorted(all_set), columns=["Контрагент"])
            df_total["1210"] = df_total["Контрагент"].map(s1210).fillna(0) / 1000
            df_total["1710"] = df_total["Контрагент"].map(s1710).fillna(0) / 1000
            df_total["3310"] = df_total["Контрагент"].map(s3310).fillna(0) / 1000
            df_total["3510"] = df_total["Контрагент"].map(s3510).fillna(0) / 1000
            df_total["общее сальдо"] = df_total["1210"] + df_total["1710"] - df_total["3310"] - df_total["3510"]
            df_total = df_total.sort_values(by="общее сальдо", ascending=False).reset_index(drop=True)
        else:
            df_total = pd.DataFrame(columns=["Контрагент", "общее сальдо"])

        out_sheet_name = safe_sheet_name(f"{prefix}сальд" if prefix else "сальд")
        if out_sheet_name in wb.sheetnames:
            wb.remove(wb[out_sheet_name])
        ws = wb.create_sheet(out_sheet_name)

        ws["A1"] = "Все значения указаны в тысячах тенге"
        ws["A1"].font = Font(name="Arial", size=10, bold=True)

        start_row = 2
        start_col = 2  # B

        font_header = Font(name="Arial", size=10, bold=True)
        font_body = Font(name="Arial", size=10)
        font_bold_body = Font(name="Arial", size=10, bold=True)
        align_center = Alignment(horizontal="center")
        align_left = Alignment(horizontal="left")
        number_format_acc = "#,##0;[Red](#,##0)"

        col_cust_contr = start_col
        col_cust_1210 = start_col + 1
        col_cust_3510 = start_col + 2
        col_cust_saldo = start_col + 3

        col_supp_contr = start_col + 5
        col_supp_1710 = start_col + 6
        col_supp_3310 = start_col + 7
        col_supp_saldo = start_col + 8

        col_total_contr = start_col + 10
        col_total_saldo = start_col + 11

        if not df_cust.empty:
            headers = {
                col_cust_contr: "Контрагент",
                col_cust_1210: "1210",
                col_cust_3510: "3510",
                col_cust_saldo: "сальдо с заказчиками",
            }
            for col, text in headers.items():
                c = ws.cell(row=start_row, column=col, value=text)
                c.font = font_header
                c.alignment = align_center

            for i, (_, row) in enumerate(df_cust.iterrows(), start=start_row + 1):
                r = i
                c_contr = ws.cell(row=r, column=col_cust_contr, value=row["Контрагент"])
                c_contr.font = font_body
                c_contr.alignment = align_left

                for col, val, style in [
                    (col_cust_1210, row["1210"], font_body),
                    (col_cust_3510, row["3510"], font_body),
                    (col_cust_saldo, row["сальдо заказчики"], font_bold_body),
                ]:
                    cell = ws.cell(row=r, column=col, value=val)
                    cell.font = style
                    cell.alignment = align_center
                    cell.number_format = number_format_acc

        if not df_supp.empty:
            headers = {
                col_supp_contr: "Контрагент",
                col_supp_1710: "1710",
                col_supp_3310: "3310",
                col_supp_saldo: "сальдо с поставщиками",
            }
            for col, text in headers.items():
                c = ws.cell(row=start_row, column=col, value=text)
                c.font = font_header
                c.alignment = align_center

            for i, (_, row) in enumerate(df_supp.iterrows(), start=start_row + 1):
                r = i
                c_contr = ws.cell(row=r, column=col_supp_contr, value=row["Контрагент"])
                c_contr.font = font_body
                c_contr.alignment = align_left

                for col, val, style in [
                    (col_supp_1710, row["1710"], font_body),
                    (col_supp_3310, row["3310"], font_body),
                    (col_supp_saldo, row["сальдо поставщики"], font_bold_body),
                ]:
                    cell = ws.cell(row=r, column=col, value=val)
                    cell.font = style
                    cell.alignment = align_center
                    cell.number_format = number_format_acc

        if not df_total.empty:
            headers = {col_total_contr: "Контрагент", col_total_saldo: "общее сальдо"}
            for col, text in headers.items():
                c = ws.cell(row=start_row, column=col, value=text)
                c.font = font_header
                c.alignment = align_center

            for i, (_, row) in enumerate(df_total.iterrows(), start=start_row + 1):
                r = i
                c_contr = ws.cell(row=r, column=col_total_contr, value=row["Контрагент"])
                c_contr.font = font_body
                c_contr.alignment = align_left

                cell = ws.cell(row=r, column=col_total_saldo, value=row["общее сальдо"])
                cell.font = font_bold_body
                cell.alignment = align_center
                cell.number_format = number_format_acc

        WIDTH_CONTR = 30
        WIDTH_NUM = 18
        for col in [col_cust_contr, col_supp_contr, col_total_contr]:
            ws.column_dimensions[get_column_letter(col)].width = WIDTH_CONTR
        for col in [
            col_cust_1210, col_cust_3510, col_cust_saldo,
            col_supp_1710, col_supp_3310, col_supp_saldo,
            col_total_saldo
        ]:
            ws.column_dimensions[get_column_letter(col)].width = WIDTH_NUM

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
# CODE 2 (Contracts) — multi-company by prefix + (md/wd) ignoring case
# Creates output sheet per prefix: "<prefix>контр" or "контр" if no prefix
# =========================
def run_code_2(file_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(file_bytes))

    prefix_to_pair: Dict[str, Dict[str, str]] = defaultdict(dict)
    for sh in wb.sheetnames:
        prefix, suf2 = split_prefix_suffix2(sh)
        if suf2 in ("wd", "md"):
            prefix = normalize_prefix(prefix)
            if suf2 not in prefix_to_pair[prefix]:
                prefix_to_pair[prefix][suf2] = sh

    valid_prefixes = [p for p, d in prefix_to_pair.items() if "wd" in d and "md" in d]
    if not valid_prefixes:
        raise ValueError("Код 2: не найдено пар листов (Wd/Md) по префиксам.")

    for prefix in valid_prefixes:
        md_ws = wb[prefix_to_pair[prefix]["md"]]
        wd_ws = wb[prefix_to_pair[prefix]["wd"]]

        payments_year = defaultdict(lambda: [0.0, 0.0, 0.0])
        performance_year = defaultdict(lambda: [0.0, 0.0, 0.0])
        payments_2025_monthly = defaultdict(lambda: [0.0] * 12)
        performance_2025_monthly = defaultdict(lambda: [0.0] * 12)

        def collect_yearly(sheet, target_dict):
            for row in range(2, sheet.max_row + 1):
                n = sheet[f"A{row}"].value
                c = sheet[f"B{row}"].value
                if not n and not c:
                    continue
                key = (str(n).strip() if n else "", str(c).strip() if c else "")
                for idx, col in enumerate(["C", "D", "E"]):
                    v = sheet[f"{col}{row}"].value
                    if v is None:
                        continue
                    try:
                        target_dict[key][idx] += float(v)
                    except:
                        pass

        def collect_monthly_2025(sheet, target_dict):
            start_col = column_index_from_string("AE")
            for row in range(2, sheet.max_row + 1):
                n = sheet[f"A{row}"].value
                c = sheet[f"B{row}"].value
                if not n and not c:
                    continue
                key = (str(n).strip() if n else "", str(c).strip() if c else "")
                for i in range(12):
                    v = sheet.cell(row=row, column=start_col + i).value
                    if v is None:
                        continue
                    try:
                        target_dict[key][i] += float(v)
                    except:
                        pass

        collect_yearly(wd_ws, payments_year)
        collect_yearly(md_ws, performance_year)
        collect_monthly_2025(wd_ws, payments_2025_monthly)
        collect_monthly_2025(md_ws, performance_2025_monthly)

        all_keys = sorted(
            set(payments_year.keys())
            | set(performance_year.keys())
            | set(payments_2025_monthly.keys())
            | set(performance_2025_monthly.keys()),
            key=lambda x: (x[0], x[1]),
        )

        out_sheet_name = safe_sheet_name(f"{prefix}контр" if prefix else "контр")
        if out_sheet_name in wb.sheetnames:
            del wb[out_sheet_name]
        ws = wb.create_sheet(out_sheet_name)

        ws["A1"] = "ИТОГО в тыс тенге"
        ws["A2"] = "Контрагент"
        ws["B2"] = "Договор"

        ws["C1"] = "оплата"
        ws["C2"] = 2023
        ws["D2"] = 2024
        ws["E2"] = 2025
        ws["F2"] = "Total"

        ws["G1"] = "выполнения с ндс"
        ws["G2"] = 2023
        ws["H2"] = 2024
        ws["I2"] = 2025
        ws["J2"] = "Total"

        ws["K2"] = "дз/(аванс)"

        ws["M1"] = "оплата"
        months = [f"2025_{str(i).zfill(2)}" for i in range(1, 13)]
        start_col_pay = column_index_from_string("M")
        for i, label in enumerate(months):
            ws[f"{get_column_letter(start_col_pay + i)}2"] = label

        ws["Y1"] = "выполнения с ндс"
        start_col_perf = column_index_from_string("Y")
        for i, label in enumerate(months):
            ws[f"{get_column_letter(start_col_perf + i)}2"] = label

        start_row = 3
        for idx, key in enumerate(all_keys):
            row = start_row + idx
            name, contract = key

            ws[f"A{row}"] = name
            ws[f"B{row}"] = contract

            py = payments_year.get(key, [0, 0, 0])
            ws[f"C{row}"] = py[0]
            ws[f"D{row}"] = py[1]
            ws[f"E{row}"] = py[2]

            pf = performance_year.get(key, [0, 0, 0])
            ws[f"G{row}"] = pf[0] * 1.12
            ws[f"H{row}"] = pf[1] * 1.12
            ws[f"I{row}"] = pf[2] * 1.12

            ws[f"F{row}"] = f"=SUM(C{row}:E{row})"
            ws[f"J{row}"] = f"=SUM(G{row}:I{row})"
            ws[f"K{row}"] = f"=J{row}-F{row}"

            mp = payments_2025_monthly.get(key, [0] * 12)
            for i in range(12):
                ws[f"{get_column_letter(start_col_pay + i)}{row}"] = mp[i]

            mf = performance_2025_monthly.get(key, [0] * 12)
            for i in range(12):
                ws[f"{get_column_letter(start_col_perf + i)}{row}"] = mf[i] * 1.12

        last_row = start_row + len(all_keys) - 1 if all_keys else 2

        regular = Font(name="Arial", size=10)
        bold = Font(name="Arial", size=10, bold=True)

        for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=column_index_from_string("AJ")):
            for c in row:
                c.font = regular

        for col in range(1, column_index_from_string("AJ") + 1):
            ws[f"{get_column_letter(col)}2"].font = bold
        for addr in ["A1", "C1", "G1", "M1", "Y1", "K2"]:
            ws[addr].font = bold

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        for col in range(1, column_index_from_string("AJ") + 1):
            addr = f"{get_column_letter(col)}2"
            ws[addr].alignment = left if addr in ["A2", "B2"] else center

        num_format = "#,##0;[Red](#,##0)"
        numeric_cols = list("CDEFGHIJK")
        numeric_cols += [get_column_letter(c) for c in range(start_col_pay, start_col_pay + 12)]
        numeric_cols += [get_column_letter(c) for c in range(start_col_perf, start_col_perf + 12)]

        for col in numeric_cols:
            for r in range(3, last_row + 1):
                cell = ws[f"{col}{r}"]
                cell.alignment = center
                cell.number_format = num_format

        for r in range(1, last_row + 1):
            ws[f"A{r}"].alignment = left
            ws[f"B{r}"].alignment = left

        for addr in ["C1", "G1", "M1", "Y1"]:
            ws[addr].alignment = left

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 38
        for col in numeric_cols:
            ws.column_dimensions[col].width = 12.2 if column_index_from_string(col) >= start_col_pay else 12.6

        thin = Side(border_style="thin", color="000000")
        border_cols = ["C", "G", "K", "M", "Y"]
        for r in range(1, last_row + 1):
            for col in border_cols:
                cell = ws[f"{col}{r}"]
                cell.border = Border(
                    left=thin,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
# UI — Theme toggle (Dark/Light), no gradients, default font
# =========================
st.set_page_config(page_title="", page_icon=None, layout="wide", initial_sidebar_state="collapsed")

if "theme" not in st.session_state:
    st.session_state.theme = "Темная"

with st.sidebar:
    st.session_state.theme = st.radio("Тема", ["Темная", "Светлая"], index=0)

# Theme variables
if st.session_state.theme == "Темная":
    BG = "#000000"
    TEXT = "#F2F2F2"
    CARD = "#0F0F0F"
    BORDER = "#2A2A2A"
    MUTED = "#BDBDBD"
    BTN_BG = "#1F1F1F"
    BTN_TEXT = "#F2F2F2"
    PROG = "#F2F2F2"
else:
    BG = "#F3EBDD"  # beige
    TEXT = "#141414"
    CARD = "#FFFFFF"
    BORDER = "#D8D0C4"
    MUTED = "#4B4B4B"
    BTN_BG = "#141414"
    BTN_TEXT = "#FFFFFF"
    PROG = "#141414"

st.markdown(
    f"""
    <style>
      #MainMenu {{visibility: hidden;}}
      footer {{visibility: hidden;}}
      header {{visibility: hidden;}}

      .stApp {{
        background: {BG};
        color: {TEXT};
      }}

      /* Keep default fonts; only set sizes and colors */
      html, body, [class*="css"], .stApp {{
        font-size: 18px !important;
      }}

      .block-container {{
        max-width: 1050px;
        padding-top: 1.8rem;
        padding-bottom: 2.0rem;
      }}

      .card {{
        background: {CARD};
        border: 1px solid {BORDER};
        border-radius: 14px;
        padding: 18px;
      }}

      .title {{
        font-size: 24px;
        font-weight: 700;
        margin: 0 0 12px 0;
        color: {TEXT};
      }}

      .sub {{
        margin: 0 0 14px 0;
        color: {MUTED};
        font-size: 16px;
      }}

      /* Inputs */
      [data-testid="stFileUploader"] section {{
        border-radius: 12px;
        padding: 12px;
        border: 1px solid {BORDER};
        background: {CARD};
      }}

      [role="radiogroup"] {{
        border-radius: 12px;
        padding: 12px 12px 8px 12px;
        border: 1px solid {BORDER};
        background: {CARD};
      }}

      /* Buttons */
      div.stButton > button {{
        width: 100%;
        border-radius: 12px !important;
        padding: 0.85rem 1rem !important;
        font-weight: 700 !important;
        font-size: 18px !important;
        background: {BTN_BG} !important;
        color: {BTN_TEXT} !important;
        border: 1px solid {BORDER} !important;
      }}

      div.stDownloadButton > button {{
        width: 100%;
        border-radius: 12px !important;
        padding: 0.85rem 1rem !important;
        font-weight: 700 !important;
        font-size: 18px !important;
        background: {BTN_BG} !important;
        color: {BTN_TEXT} !important;
        border: 1px solid {BORDER} !important;
      }}

      /* Progress bar */
      div[data-testid="stProgress"] > div > div {{
        background-color: {PROG} !important;
      }}

      /* Text colors for markdown */
      .stMarkdown, .stMarkdown p, .stCaption, label {{
        color: {TEXT} !important;
      }}
      .stCaption {{
        color: {MUTED} !important;
      }}
    </style>
    """,
    unsafe_allow_html=True,
)

# Simple header (no gradients)
st.markdown(f"<div class='card'><div class='title'>Обработка Excel</div><div class='sub'>Загрузите файл, выберите сценарий и скачайте результат.</div></div>", unsafe_allow_html=True)
st.write("")

left, right = st.columns([1.05, 0.95], gap="large")

with left:
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("<div class='title'>Файл</div>", unsafe_allow_html=True)
    st.markdown("<div class='sub'>Поддерживаются форматы .xlsx и .xlsm</div>", unsafe_allow_html=True)
    uploaded = st.file_uploader("Загрузите Excel файл", type=["xlsx", "xlsm"], label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)

    st.write("")

    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("<div class='title'>Сценарий</div>", unsafe_allow_html=True)
    st.markdown("<div class='sub'>Выберите нужную обработку</div>", unsafe_allow_html=True)
    mode = st.radio(
        "Выберите обработку",
        options=["Сальдо", "Контракты", "Оба (Сальдо → Контракты)"],
        index=0,
        label_visibility="collapsed",
    )
    st.markdown("</div>", unsafe_allow_html=True)

with right:
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("<div class='title'>Запуск</div>", unsafe_allow_html=True)
    st.markdown("<div class='sub'>Нажмите «Обработать». После завершения появится кнопка скачивания.</div>", unsafe_allow_html=True)

    run_btn = st.button("Обработать", disabled=(uploaded is None))

    status_box = st.empty()
    progress = st.progress(0)

    if uploaded is not None:
        size_mb = len(uploaded.getvalue()) / (1024 * 1024)
        st.markdown(f"<div class='sub'>Файл: <b>{uploaded.name}</b><br/>Размер: <b>{size_mb:.2f} MB</b></div>", unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)

st.write("")

if run_btn and uploaded is not None:
    file_bytes = uploaded.getvalue()

    try:
        status_box.info("Подготовка…")
        progress.progress(12)
        time.sleep(0.10)

        out_bytes = file_bytes

        if mode in ["Сальдо", "Оба (Сальдо → Контракты)"]:
            status_box.info("Обработка: Сальдо…")
            progress.progress(40)
            out_bytes = run_code_1(out_bytes)
            progress.progress(64)

        if mode in ["Контракты", "Оба (Сальдо → Контракты)"]:
            status_box.info("Обработка: Контракты…")
            progress.progress(78)
            out_bytes = run_code_2(out_bytes)
            progress.progress(92)

        status_box.success("Готово. Можно скачать результат.")
        progress.progress(100)

        st.markdown("<div class='card'>", unsafe_allow_html=True)
        st.markdown("<div class='title'>Результат</div>", unsafe_allow_html=True)

        st.download_button(
            label="Скачать обработанный Excel",
            data=out_bytes,
            file_name=f"processed_{uploaded.name}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.markdown("</div>", unsafe_allow_html=True)

    except Exception as e:
        progress.progress(0)
        status_box.error(f"Ошибка: {e}")
